import os
import openpyxl

xl_path = "fonts.xlsx"
font_single = "project_fonts/Single"
font_diversity = "project_fonts/Diversity"

single_font_name = os.listdir(font_single)
diversity_font_dir = os.listdir(font_diversity)

font_xl = openpyxl.load_workbook(xl_path)

font_xl.remove(font_xl['name'])
font_xl.create_sheet(title='name')

s1 = font_xl['name']
s1.append(['Number', 'Category', 'File name', 'Font name'])

for index, name in enumerate(single_font_name):
    s1.cell(index+2, 1).value = index + 1
    s1.cell(index+2, 2).value = index
    s1.cell(index+2, 3).value = name
    s1.cell(index+2, 4).value = name.split('.')[0]

category = number = index + 1

for name in diversity_font_dir:
    i = font_diversity + "/" + name
    sub_font = os.listdir(i)
    for index, font in enumerate(sub_font):
        s1.cell((number+2 + index), 1).value = number + index + 1
        s1.cell((number+2 + index), 2).value = category
        s1.cell((number+2 + index), 3).value = font
        s1.cell((number+2 + index), 4).value = name
    number = number + 1 + index
    category += 1

font_xl.save(xl_path)